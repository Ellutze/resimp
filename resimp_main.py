# -*- coding: utf-8 -*-
"""
Created on Thu Feb 17 14:50:17 2022

@author: jakub.kucera
"""

#layout:
# import file field, PAM-RTM install location, initiate button



#accomodate for file formats (default the excel from Cristian, 2nd singular csv table (3rows, repeated?)(3rows with various temp in same col?)

import shutil
#import sys
import time
#for command line use:
import subprocess
#for excel imports:
import openpyxl
#for general mathematical operations
import numpy as np

import os


def resimp(rnm,fn,esi):
    #esi is the location of PAM-RTM installation
    #rnm is the resin name
    #fn is the source file for data (either csv or xlsx)
    lPath = os.getcwd()
    if ".xlsx" in fn:
        
        #print(lPath)
        # Give the location of the file
        fll = fn  # replace this with the above
        #fn = "filenametest" # replace with the UI input
         
        #workbook open
        wb_obj = openpyxl.load_workbook(fll,data_only = True)
        #active sheet (make sure it is always on the correct sheet?)
        sheet_obj = wb_obj.active
        
        r = 14
        c = 2 
        e = True
        i = 0
        Ts = []
        while e == True:
            cell_obj = sheet_obj.cell(row = r, column = c)
            
            try:
                if i > 0 and int(cell_obj.value) <= 0:
                    e = False
                else:
                    t = float(cell_obj.value) + 273.15 # assuming source always in C? what if in Kelvin
                    #print(t)
                    t = str(t)
                    Ts.append(t)
                    c = c + 7
                    i = i + 1    
            
            #if i > 0 and int(cell_obj.value) <= 0:
            except: 
                # Print value of cell object
                # using the value attribute
                e = False
        #print(Ts)
        i = 0
        #M = np.zeros([1,2])
        for t in Ts:
            r = 19
            c = 3 + 7*i
            with open(lPath+'\\tt\\'+str(t)+'.txt', 'w') as f:
                e = True
                while e == True:
                    a = sheet_obj.cell(row = r, column = c)
                    dadt = sheet_obj.cell(row = r, column = c+3)
                
                    if a.value != None:
                        a = format(a.value,'.30f')
                        dadt = format(dadt.value,'.30f')
                        f.write(str(a)+" "+str(dadt)+"\n")
                        #m_temp = np.matriix([a,dadt])
                        #M = np.concatenate((M,m_temp),axis=0)
                        r = r + 1
                    else:
                        e = False
            #delete the 0 row 
            
            #save matrix as txt file
            
            i = i + 1
    elif ".csv" in fn:
        
        file = open(fn)
        raw_inp = np.loadtxt(file, delimiter=",")
        temps = int(np.size(raw_inp,1)/3)
        e = np.size(raw_inp,0)
        i = 0 
        Ts = []
        while i < temps:
            
            t = raw_inp[0,3*i] + 273.15 # assuming source always in C
            Ts.append(t)
            with open(lPath+'\\tt\\'+str(t)+'.txt', 'w') as f:
                ii = 0
                while ii < e:
                    a = raw_inp[ii,3*i+1]
                    
                    dadt = raw_inp[ii,3*i+2]
        
                    a = format(a,'.30f')
                    dadt = format(dadt,'.30f')
                    f.write(str(a)+" "+str(dadt)+"\n")
                    #m_temp = np.matriix([a,dadt])
                    #M = np.concatenate((M,m_temp),axis=0)
                    ii = ii +1
            i = i + 1
    #outputs to be picked up by the PAM-RTM script
    np.save(lPath+"\\temp_list.npy",Ts,allow_pickle=True, fix_imports=True)
    with open(lPath+"\\filename.txt",'w') as f:
        f.write(rnm)
    
    
    
    #save file with meta-data  
    
    #initiate command line 
        
    #REPLACE FILE -- RELATIVE LOCATION TO THIS SCRIPT
      
    #moving PAM-RTM internal script reference location
    with open(lPath+"\\imp_PR.py",'r') as f:
        ff = str(f.read())
        #print(ff)
        
    with open(lPath+"\\imp_PR.py",'w') as f:   
        ins = """lPath='"""+lPath+"""'#"""
        ins = ins.replace("""\\""","""\\\\""")
        fff = ff.replace("""lPath=""",ins)
        f.write(fff)
        
    def cmd2(command,RTMFile=""):
        #Passes command to command line.
        #Required due to Visual-RTM having it's own library of python packages.
        process = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True, cwd=esi) #"C:\\Program Files\\ESI Group\\Visual-Environment\\17.0\\Windows-x64")
        proc_stdout = process.communicate()#[0].strip()
        #print(command)
        #print(proc_stdout)
        print("Your material is being imported...")
        
    cmd2("VEBatch -activeconfig Trade:CompositesandPlastics -activeapp VisualRTM -sessionrun "+lPath+"\\imp_PR.py")
    
    
    #after import is done make sure to delete everything in "tt" (temporary temperature tables)
    #not sure if I need to give time?
    time.sleep(30)
    ff = lPath+'\\tt'
    for filename in os.listdir(ff):
        file_path = os.path.join(ff, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))
